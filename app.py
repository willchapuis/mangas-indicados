from flask import Flask, render_template, request, redirect, url_for
import os
import json
from openpyxl import load_workbook
from rapidfuzz import fuzz
from werkzeug.utils import secure_filename

app = Flask(__name__)

os.makedirs("uploads", exist_ok=True)
app.config['UPLOAD_FOLDER'] = 'uploads'

DADOS_PATH = "obras.json"

# Carrega ou inicia os dados
def carregar_dados():
    if not os.path.exists(DADOS_PATH):
        with open(DADOS_PATH, "w", encoding="utf-8") as f:
            json.dump({}, f, ensure_ascii=False, indent=2)
    try:
        with open(DADOS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError:
        with open(DADOS_PATH, "w", encoding="utf-8") as f:
            json.dump({}, f, ensure_ascii=False, indent=2)
        return {}

def salvar_dados(dados):
    with open(DADOS_PATH, 'w', encoding='utf-8') as f:
        json.dump(dados, f, indent=2, ensure_ascii=False)

def normalizar_nome(nome):
    nome = nome.lower().strip()
    nome = nome.replace("***", "").replace("...", "").replace("---", "")
    nome = nome.replace("?", "").replace("!", "").replace(",", "")
    nome = nome.replace("99+", "+99")  # exemplo específico
    nome = ' '.join(nome.split())  # remove espaços duplos
    return nome

# Adiciona nova indicação
def adicionar_indicacao(lista_obras, dados):
    for obra in lista_obras:
        obra_original = obra.strip()
        if not obra_original:
            continue
        obra_normalizada = normalizar_nome(obra_original)

        similar = None
        for existente in dados:
            if fuzz.ratio(obra_normalizada, normalizar_nome(existente)) >= 90:
                similar = existente
                break

        if similar:
            dados[similar]['indicacoes'] += 1
        else:
            dados[obra_original] = {
                'indicacoes': 1
            }
    salvar_dados(dados)

# Rota principal
@app.route('/')
def index():
    dados = carregar_dados()
    obras = sorted(dados.items(), key=lambda x: -x[1]['indicacoes'])
    maior = obras[0][1]['indicacoes'] if obras else 1
    return render_template('index.html', obras=obras, maior=maior)

# Upload de .xlsx
@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        file = request.files['file']
        if file.filename.endswith('.xlsx'):
            filename = secure_filename(file.filename)
            path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(path)

            wb = load_workbook(path)
            ws = wb.active

            blocos = []
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                valor = row[0]
                if valor is None:
                    if blocos:
                        adicionar_indicacao(blocos, carregar_dados())
                        blocos = []
                else:
                    blocos.append(str(valor))

            if blocos:
                adicionar_indicacao(blocos, carregar_dados())

            return redirect(url_for('index'))
    return render_template('upload.html')

# Adição manual
@app.route('/adicionar', methods=['POST'])
def adicionar():
    texto = request.form.get('novas_obras', '')
    lista_obras = texto.strip().splitlines()
    adicionar_indicacao(lista_obras, carregar_dados())
    return redirect(url_for('index'))

# Remover obra
@app.route('/excluir/<obra>', methods=['POST'])
def excluir(obra):
    dados = carregar_dados()
    if obra in dados:
        del dados[obra]
        salvar_dados(dados)
    return redirect(url_for('index'))

if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)
    app.run(debug=True)
